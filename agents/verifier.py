"""
Agent 3: Verifier

Reads the generated Excel file back using openpyxl and performs
Python-level cross-checks (not relying solely on Excel formula results).

Checks:
1. Balance Sheet equation: Total Assets ≈ Total Liabilities + Equity
2. Cash Flow reconciliation: Net Change ≈ CFO + CFI + CFF
3. PAT consistency: IS PAT ≈ CF Net Income (within 5%)
4. No formula error strings in any cell (#DIV/0!, #REF!, #VALUE!)
"""
from typing import Optional

import openpyxl

from models.company_data import FinancialData
from models.extraction_result import VerificationReport
from utils.logger import get_logger

log = get_logger()

_ERROR_STRINGS = {"#DIV/0!", "#REF!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def _safe_float(cell) -> Optional[float]:
    """Read a numeric cell value; return None if empty or non-numeric."""
    if cell is None or cell.value is None:
        return None
    try:
        return float(cell.value)
    except (TypeError, ValueError):
        val = str(cell.value)
        if any(e in val for e in _ERROR_STRINGS):
            return None
        return None


def _check_formula_errors(ws, report: VerificationReport):
    """Scan a worksheet for Excel error strings."""
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and any(e in val for e in _ERROR_STRINGS):
                report.add_issue(
                    f"Formula error in {ws.title}!{cell.coordinate}: {val}"
                )


def verify_workbook(excel_path: str, financial_data: FinancialData) -> VerificationReport:
    """
    Read the Excel file and run all cross-checks.
    Returns a VerificationReport.
    """
    report = VerificationReport()
    log.info(f"Verifying workbook: {excel_path}")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        report.add_issue(f"Could not open workbook: {e}")
        return report

    years = financial_data.sorted_years()

    # ── 1. Check for formula error strings in all sheets ──────────────────────
    for ws in wb.worksheets:
        _check_formula_errors(ws, report)

    # ── 2. Python-level BS equation check ────────────────────────────────────
    if "Balance Sheet" in wb.sheetnames:
        bs_ws = wb["Balance Sheet"]
        from agents.excel_builder import BS, _data_cols
        data_cols = _data_cols(years)

        for i, yr in enumerate(years):
            col_idx = data_cols[i] + 1  # openpyxl is 1-based
            ta = _safe_float(bs_ws.cell(row=BS["total_assets"], column=col_idx))
            tle = _safe_float(bs_ws.cell(row=BS["total_le"], column=col_idx))

            if ta is not None and tle is not None:
                diff = abs(ta - tle)
                if diff > 1.0:
                    report.add_issue(
                        f"Balance Sheet {yr}: Assets={ta:,.2f} ≠ L+E={tle:,.2f} "
                        f"(diff={diff:,.2f})"
                    )
                else:
                    log.info(f"  BS check {yr}: PASS (diff={diff:.4f})")
            else:
                report.add_warning(f"BS {yr}: Could not read Total Assets or L+E (may be blank)")

    # ── 3. Python-level CF reconciliation (from Pydantic source data) ────────
    # Note: openpyxl reads formula cells as 0 (no cached values in new files),
    # so we use the original FinancialData models for this check instead.
    for yr in years:
        cf = financial_data.get_cash_flow(yr)
        if cf is None:
            continue
        cfo = cf.cash_from_operations
        cfi = cf.cash_from_investing
        cff = cf.cash_from_financing
        net = cf.net_change_in_cash

        if all(v is not None for v in [cfo, cfi, cff, net]):
            computed = cfo + cfi + cff
            if abs(computed) > 0.01:
                diff_pct = abs(net - computed) / abs(computed)
                if diff_pct > 0.02:
                    report.add_issue(
                        f"CF reconciliation {yr}: CFO+CFI+CFF={computed:,.2f} "
                        f"but net_change={net:,.2f} ({diff_pct:.1%} diff)"
                    )
                else:
                    log.info(f"  CF check {yr}: PASS")
        elif all(v is not None for v in [cfo, cfi, cff]):
            log.info(f"  CF check {yr}: net_change not provided — skipping reconciliation")

    # ── 4. PAT consistency check (from Pydantic source data) ─────────────────
    for yr in years:
        is_stmt = financial_data.get_income_statement(yr)
        cf_stmt = financial_data.get_cash_flow(yr)
        if is_stmt is None or cf_stmt is None:
            continue
        pat = is_stmt.pat
        net_inc = cf_stmt.net_income
        if pat is not None and net_inc is not None and abs(pat) > 0.01:
            diff_pct = abs(pat - net_inc) / abs(pat)
            if diff_pct > 0.05:
                report.add_warning(
                    f"PAT consistency {yr}: IS PAT={pat:,.2f}, CF Net Income={net_inc:,.2f} "
                    f"({diff_pct:.1%} difference — may be due to minority interest)"
                )
            else:
                log.info(f"  PAT check {yr}: PASS")

    # ── Summary ───────────────────────────────────────────────────────────────
    if report.passes:
        log.info(f"Verification: {report.summary()}")
    else:
        log.warning(f"Verification: {report.summary()}")
        for issue in report.issues:
            log.warning(f"  ✗ {issue}")
    for warn in report.warnings:
        log.warning(f"  ⚠ {warn}")

    return report
